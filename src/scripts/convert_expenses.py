#!/usr/bin/env python3
"""
Script to convert Excel expense data to JSON format.
Usage: python convert_expenses.py <path_to_excel_file.xlsx>
"""
import openpyxl
import json
import sys
from pathlib import Path

def safe_int(value):
    """Safely convert value to int, return None if not possible."""
    if value is None:
        return None
    try:
        # Handle float values that are actually integers (e.g., 10.0)
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return int(value)
    except (ValueError, TypeError):
        return None

def safe_str(value):
    """Safely convert value to string, return None if empty."""
    if value is None:
        return None
    str_value = str(value).strip()
    return str_value if str_value else None

def convert_excel_to_json(excel_path):
    """Convert Excel file to JSON format."""
    # Validate file exists
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    # Validate file extension
    if excel_path.suffix.lower() != '.xlsx':
        print(f"Error: File must have .xlsx extension, got: {excel_path.suffix}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)
    
    # Get the "podział limitów" sheet
    if "podział limitów" not in workbook.sheetnames:
        print(f"Error: Sheet 'podział limitów' not found in workbook")
        print(f"Available sheets: {workbook.sheetnames}")
        sys.exit(1)
    
    sheet = workbook["podział limitów"]
    
    expenses = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        # Extract required fields (using 0-based indexing)
        chapter = safe_int(row[7]) if len(row) > 7 else None  # Column H (index 7)
        task_name = safe_str(row[11]) if len(row) > 11 else None  # Column L (index 11)
        financial_needs = safe_int(row[14]) if len(row) > 14 else None  # Column O (index 14)
        
        # Skip if required fields are missing
        if not chapter or not task_name or not financial_needs or financial_needs <= 0:
            continue
        
        # Extract all additional fields
        expense_data = {
            "chapter": chapter,
            "task_name": task_name,
            "financial_needs": financial_needs,
            # Additional fields from Excel
            "czesc": safe_int(row[0]) if len(row) > 0 else None,
            "departament": safe_str(row[1]) if len(row) > 1 else None,
            "rodzaj_projektu": safe_str(row[2]) if len(row) > 2 else None,
            "opis_projektu": safe_str(row[3]) if len(row) > 3 else None,
            "data_zlozenia": safe_str(row[4]) if len(row) > 4 else None,
            "program_operacyjny": safe_str(row[5]) if len(row) > 5 else None,
            "termin_realizacji": safe_str(row[6]) if len(row) > 6 else None,
            "zrodlo_fin": safe_int(row[8]) if len(row) > 8 else None,
            "bz": safe_str(row[9]) if len(row) > 9 else None,
            "beneficjent": safe_str(row[10]) if len(row) > 10 else None,
            "szczegolowe_uzasadnienie": safe_str(row[12]) if len(row) > 12 else None,
            "budget_2025": safe_int(row[13]) if len(row) > 13 else None,
            "budget_2026": safe_int(row[14]) if len(row) > 14 else None,
            "budget_2027": safe_int(row[15]) if len(row) > 15 else None,
            "budget_2028": safe_int(row[16]) if len(row) > 16 else None,
            "budget_2029": safe_int(row[17]) if len(row) > 17 else None,
            "etap_dzialan": safe_str(row[18]) if len(row) > 18 else None,
            "umowy": safe_str(row[19]) if len(row) > 19 else None,
            "nr_umowy": safe_str(row[20]) if len(row) > 20 else None,
            "z_kim_zawarta": safe_str(row[21]) if len(row) > 21 else None,
            "uwagi": safe_str(row[22]) if len(row) > 22 else None,
        }
        
        expenses.append(expense_data)
    
    # Determine output path
    output_path = Path(__file__).parent.parent / "data" / "expenses_template.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save to JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(expenses, f, ensure_ascii=False, indent=2)
    
    print(f"Successfully converted {len(expenses)} expenses")
    print(f"Saved to: {output_path}")
    
    return len(expenses)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python convert_expenses.py <path_to_excel_file.xlsx>")
        print("Example: python convert_expenses.py docs/expenses.xlsx")
        sys.exit(1)
    
    excel_file = Path(sys.argv[1])
    convert_excel_to_json(excel_file)
